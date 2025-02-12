import  openpyxl
# File ->workbook-->Sheets-->Rows-->Cells

file=r"C:\Users\Alian\Desktop\YM\data.xlsx" #path of file
wb= openpyxl.load_workbook(file) #to load workbook
sheet = wb["Sheet1"] #Sheet of data

row = sheet.max_row #count number of rows in Excel sheet
col = sheet.max_column #count number of columns in Excel sheet

#Reading all the rows & columns from Excel Sheet
for r in range(1,row+1):
    for c in range(1,col+1):
        print(sheet.cell(r,c).value, end='        ') #to read from cell
    print()


