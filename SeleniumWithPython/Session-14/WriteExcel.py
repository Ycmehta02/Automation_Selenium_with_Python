import openpyxl

## Writes Same Data--------------------------------------------
# file =r"C:\Users\Alian\Desktop\YM\data1.xlsx" #paTH TO FILE
# wb = openpyxl.load_workbook(file) #load the workbook
# # sheet = wb["Sheet1"] # can also write sheet name
# sheet = wb.active #loads active sheet (single sheet)
# for r in range(1,6):
#     for c in range(1,4): #5 rows & 3 columns
#         sheet.cell(r,c).value="Hello" #to write in cell
#
# wb.save(file)

## Writes Multiple Data-----------------------------------------
file =r"C:\Users\Alian\Desktop\YM\data1.xlsx" #paTH TO FILE
wb = openpyxl.load_workbook(file) #load the workbook
# sheet = wb["Sheet1"] # can also write sheet name
sheet = wb.active #loads active sheet (single sheet)

sheet.cell(1,1).value=123
sheet.cell(1,2).value="YM"
sheet.cell(1,3).value="QA"

sheet.cell(2,1).value=134
sheet.cell(2,2).value="HB"
sheet.cell(2,3).value="QA"

sheet.cell(3,1).value=156
sheet.cell(3,2).value="SP"
sheet.cell(3,3).value="QA"

wb.save(file) #save the file after entering the data









